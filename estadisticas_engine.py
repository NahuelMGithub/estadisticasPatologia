import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import tempfile


BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "data" / "template_estadisticas.xlsx"


def procesar_expedientes(expedientes_file, mes_seleccionado):
    # -----------------------
    # LEER EXCEL
    # -----------------------
    df = pd.read_excel(expedientes_file)

    # Normalizar columnas
    df["DEPTO. JUDICIAL"] = df["DEPTO. JUDICIAL"].astype(str).str.upper()
    df["FECHA DE SOLICITUD"] = pd.to_datetime(df["FECHA DE SOLICITUD"], errors="coerce")

    # Filtrar por mes
    mes = int(mes_seleccionado)
    df = df[df["FECHA DE SOLICITUD"].dt.month == mes]

    # -----------------------
    # CARGAR TEMPLATE
    # -----------------------
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active  # Hoja 1

    # -----------------------
    # MAPEO (A6:A25 -> R6:R25)
    # -----------------------
    for row in range(6, 26):
        depto = ws[f"A{row}"].value

        if depto:
            depto = str(depto).upper()

            count = df[df["DEPTO. JUDICIAL"] == depto].shape[0]

            ws[f"R{row}"] = count

    # -----------------------
    # GUARDAR TEMPORAL
    # -----------------------
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)

    return tmp_file.name