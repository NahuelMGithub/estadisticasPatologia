print("🔥 ESTOY EJECUTANDO ESTE APP.PY")

from flask import Flask, request, send_file, jsonify, render_template
from flask_cors import CORS
import pandas as pd
from io import BytesIO, StringIO
import traceback
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path

app = Flask(__name__)
CORS(app)


# -------------------------
# 🌐 FRONTEND
# -------------------------
@app.route("/")
def home():
    return render_template("index.html")


# -------------------------
# 🧪 TEST
# -------------------------
@app.route("/ping")
def ping():
    return "OK"


# -------------------------
# 🧠 LECTURA INTELIGENTE
# -------------------------
def leer_archivo(file, filename):
    filename = filename.lower()

    try:
        content = file.read()
        file.seek(0)

        contenido_lower = content.lower()
        es_html = (
            b"<html" in contenido_lower
            or b"<table" in contenido_lower
            or b"<style" in contenido_lower
            or b"<!doctype" in contenido_lower
        )

        if es_html:
            html_str = None
            for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
                try:
                    html_str = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    pass

            if html_str is None:
                html_str = content.decode("latin-1", errors="replace")

            table_pos = html_str.lower().find("<table")
            if table_pos == -1:
                raise ValueError("No se encontraron tablas HTML en el archivo subido")

            html_str = html_str[table_pos:]
            tables = pd.read_html(StringIO(html_str), flavor="html5lib")
            return tables[0]

        elif filename.endswith(".xlsx"):
            return pd.read_excel(BytesIO(content), engine="openpyxl")

        elif filename.endswith(".xls"):
            return pd.read_excel(BytesIO(content), engine="xlrd")

        else:
            raise Exception("Formato no soportado")

    except Exception as e:
        print("❌ ERROR LEYENDO ARCHIVO:", str(e))
        raise

# -------------------------
# 🔧 LÓGICA TRANSFORMAR
# -------------------------
def separar_ufi(df):
    df = df.copy()
    i = len(df) - 1

    while i > 0:
        try:
            actual = str(df.iloc[i, 6]).upper()
            anterior = str(df.iloc[i - 1, 6]).upper()

            if "UNIDAD FUNCIONAL" in actual and "MORGUE" in anterior:
                df.iloc[i - 1, 7] = df.iloc[i, 6]
                df = df.drop(df.index[i]).reset_index(drop=True)
                i -= 1

        except Exception as e:
            print("⚠️ Error:", e)

        i -= 1

    return df


# -------------------------
# 🚀 TRANSFORMAR
# -------------------------
@app.route("/transformar", methods=["POST"])
def transformar():
    try:
        file = request.files["file"]
        filename = file.filename

        df = leer_archivo(file, filename)
        df = separar_ufi(df)

        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        return send_file(output, download_name="transformado.xlsx", as_attachment=True)

    except Exception as e:
        return jsonify({"error": str(e), "detalle": traceback.format_exc()}), 500


# -------------------------
# 📊 ESTADÍSTICAS
# -------------------------
@app.route("/estadisticas", methods=["POST"])
def estadisticas():
    try:
        file = request.files["pericias"]
        mes = int(request.form.get("mes"))
        anio = int(request.form.get("anio"))  # 👈 ESTO FALTABA
        df = pd.read_excel(file, sheet_name="Hoja 1")

        columnas_requeridas = ["INGRESO", "DPTO.", "MATERIAL RECIBIDO", "TIPO DE MUESTRA ROTULO", "Columna1"]
        faltantes = [col for col in columnas_requeridas if col not in df.columns]
        if faltantes:
            raise ValueError(f"Faltan columnas requeridas en Pericias: {', '.join(faltantes)}")

        df["DPTO."] = df["DPTO."].astype(str).str.upper()
        df["INGRESO"] = pd.to_datetime(df["INGRESO"], errors="coerce")
        df["MATERIAL RECIBIDO"] = df["MATERIAL RECIBIDO"].fillna("").astype(str).str.upper()
        df["MUESTRA_NORMALIZADA"] = (
            df["TIPO DE MUESTRA ROTULO"].fillna("").astype(str)
            + " "
            + df["Columna1"].fillna("").astype(str)
        ).str.upper()

        df = df[
    (df["INGRESO"].dt.month == mes) &
    (df["INGRESO"].dt.year == anio)
]

        # -------------------------
        # TEMPLATE
        # -------------------------
        BASE_DIR = Path(__file__).resolve().parent
        template_path = BASE_DIR / "data" / "template_estadisticas.xlsx"

        wb = load_workbook(template_path)
        ws = wb["Hoja1"]

        # -------------------------
        # MAPEO A6:A25 → B6:P25
        # -------------------------
        for row in range(6, 26):
            depto = ws[f"A{row}"].value

            if depto:
                depto = str(depto).upper()
                df_depto = df[df["DPTO."] == depto]
                material = df_depto["MATERIAL RECIBIDO"]
                muestra = df_depto["MUESTRA_NORMALIZADA"]
                frascos = pd.to_numeric(
                    material.str.extract(r"(\d+)\s*FP", expand=False),
                    errors="coerce"
                ).fillna(0)

                ws[f"B{row}"] = df_depto.shape[0]
                ws[f"C{row}"] = int(frascos.sum())
                ws[f"D{row}"] = ((material != "") & ~material.str.contains("FP", na=False)).sum()
                ws[f"E{row}"] = (muestra.str.contains("FETO", na=False) & ~muestra.str.contains("POOL FETO", na=False)).sum()
                ws[f"F{row}"] = muestra.str.contains("PLACENTA", na=False).sum()
                ws[f"G{row}"] = muestra.str.contains("POOL ADULTO", na=False).sum()
                ws[f"H{row}"] = muestra.str.contains("POOL MENOR", na=False).sum()
                ws[f"I{row}"] = muestra.str.contains("POOL FETO", na=False).sum()
                ws[f"J{row}"] = (
                    muestra.str.contains("VÍA AÉREA", na=False)
                    | muestra.str.contains("VIA AEREA", na=False)
                    | muestra.str.contains("VIA AREA", na=False)
                    | muestra.str.contains("VÍA AREA", na=False)
                ).sum()
                ws[f"K{row}"] = muestra.str.contains("HAF", na=False).sum()
                ws[f"L{row}"] = muestra.str.contains("HAB", na=False).sum()
                ws[f"M{row}"] = muestra.str.contains("ELECTROCU", na=False).sum()
                ws[f"N{row}"] = muestra.str.contains("SURCO", na=False).sum()
                ws[f"O{row}"] = (
                    muestra.str.contains("LOSANGE S/E", na=False)
                    | (muestra.str.contains("LOSANGE", na=False) & muestra.str.contains("S/E", na=False))
                ).sum()
                ws[f"P{row}"] = muestra.str.contains("SIN ESPECIFICAR", na=False).sum()

                if depto == "LA PLATA":
                    diagnostico = {
                        "B6 total": int(ws[f"B{row}"].value),
                        "C6 frascos": int(ws[f"C{row}"].value),
                        "D6 otros": int(ws[f"D{row}"].value),
                        "E6 feto": int(ws[f"E{row}"].value),
                        "F6 placenta": int(ws[f"F{row}"].value),
                        "G6 pool adulto": int(ws[f"G{row}"].value),
                        "H6 pool menor": int(ws[f"H{row}"].value),
                        "I6 pool feto": int(ws[f"I{row}"].value),
                        "J6 via aerea": int(ws[f"J{row}"].value),
                        "K6 HAF": int(ws[f"K{row}"].value),
                        "L6 HAB": int(ws[f"L{row}"].value),
                        "M6 electrocucion": int(ws[f"M{row}"].value),
                        "N6 surco": int(ws[f"N{row}"].value),
                        "O6 losange sin especificar": int(ws[f"O{row}"].value),
                        "P6 sin especificar": int(ws[f"P{row}"].value),
                    }
                    print("DIAGNOSTICO LA PLATA")
                    print("Registros filtrados:", df_depto.shape[0])
                    print("MATERIAL RECIBIDO unicos:", sorted(material.unique()))
                    print("MUESTRA_NORMALIZADA unicos:", sorted(muestra.unique()))
                    print("Resultados B6:P6:", diagnostico)

        # -------------------------
        # 🔢 SUMA B26:R26
        # -------------------------
        for col in range(2, 19):
            col_letter = ws.cell(row=6, column=col).column_letter

            total = 0
            for row in range(6, 26):
                value = ws[f"{col_letter}{row}"].value
                try:
                    total += float(value)
                except:
                    pass

            ws[f"{col_letter}26"] = total

        # -------------------------
        # OUTPUT
        # -------------------------
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        meses = {
            1: "ENERO", 2: "FEBRERO", 3: "MARZO",
            4: "ABRIL", 5: "MAYO", 6: "JUNIO",
            7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE",
            10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
        }

        nombre_mes = meses.get(mes, "MES")
        año = anio

        return send_file(
            output,
            download_name=f"Estadísticas {nombre_mes} {año}.xlsx",
            as_attachment=True
        )

    except Exception as e:
        print("❌ ERROR ESTADÍSTICAS:")
        print(traceback.format_exc())

        return jsonify({
            "error": str(e),
            "detalle": traceback.format_exc()
        }), 500


# -------------------------
# ▶️ RUN
# -------------------------
if __name__ == "__main__":
    app.run(
        debug=True,
        host="127.0.0.1",
        port=8000
    )