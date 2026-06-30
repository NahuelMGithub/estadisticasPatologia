import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# -------------------------
# CONFIG
# -------------------------
COLUMN_DEPARTAMENTO = "DEPTO. JUDICIAL"
COLUMN_FECHA = "FECHA DE SOLICITUD"

MESES = {
    "1": 1, "2": 2, "3": 3, "4": 4,
    "5": 5, "6": 6, "7": 7, "8": 8,
    "9": 9, "10": 10, "11": 11, "12": 12
}

# -------------------------
# PROCESAR EXPEDIENTES
# -------------------------
def procesar_expedientes(excel_path, template_path, mes_seleccionado, output_path):

    # Leer datos
    df = pd.read_excel(excel_path)

    # Normalizar
    df[COLUMN_DEPARTAMENTO] = df[COLUMN_DEPARTAMENTO].astype(str).str.upper()
    df[COLUMN_FECHA] = pd.to_datetime(df[COLUMN_FECHA], errors="coerce")

    mes = int(mes_seleccionado)

    # Filtrar por mes
    df_filtrado = df[df[COLUMN_FECHA].dt.month == mes]

    # Cargar template
    wb = load_workbook(template_path)
    ws = wb.active  # Hoja 1

    # Recorrer filas del Excel modelo (A6:A25)
    for row in range(6, 26):
        depto = ws[f"A{row}"].value

        if depto:
            depto = str(depto).upper()

            count = df_filtrado[
                df_filtrado[COLUMN_DEPARTAMENTO] == depto
            ].shape[0]

            ws[f"R{row}"] = count

    # Guardar resultado
    wb.save(output_path)